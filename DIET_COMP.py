# ficomp_dietcomp_creator.py
import streamlit as st
import pandas as pd
import os
import re
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

st.set_page_config(page_title="🐄 FICOMP & DIETCOMP Creator", layout="wide")
st.title("🐄 Create FICOMP & DIETCOMP Sheets from CVAS Excel")

# ----------------------------
# Helpers
# ----------------------------
def norm_text(s: str) -> str:
    s = str(s).lower()
    s = s.replace("\u00a0", " ")
    s = re.sub(r"[_\-]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()

@st.cache_data
def load_nasem_library():
    nasem_path = os.path.join(os.path.dirname(__file__), "NASEM_2021_FEED_LIBRARY.csv")
    nasem_df = pd.read_csv(nasem_path)
    nasem_df["Feed Name"] = nasem_df["Feed Name"].astype(str)
    nasem_df["Feed Name (Original)"] = nasem_df["Feed Name"]
    nasem_df["Feed Name (Clean)"] = nasem_df["Feed Name"].apply(norm_text)
    return nasem_df

nasem_df_builtin = load_nasem_library()

def auto_fit_and_format(output_file):
    wb = load_workbook(output_file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Wrap text in Key sheet
        if sheet == "KEY":
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 2
    wb.save(output_file)

# ----------------------------
# CVAS Template
# ----------------------------
st.subheader("📥 CVAS Reference Template")
example_cvas = pd.DataFrame({
    "desc_1": ["CORN SILAGE SHORT", "CORN GRAIN", "SOYBEAN MEAL"],
    "feedtype": ["FORAGE", "GRAIN", "PROTEIN"],
    "NDF": [37.2, 9.5, 12.0],
    "aNDFom": [35.3, 9.5, 11.8],
    "RDP": [16.5, 5.0, 30.0],
    "CP": [24.4, 8.6, 48.0],
    "TFA": [2.61, 3.64, 1.0],
    "Ash": [3.66, 1.35, 6.0],
    "ADF": [22.0, 2.8, 10.0],
    "Lignin": [2.72, 1.69, 0.5],
    "Starch": [35.5, 75.0, 2.0],
    "NDFD48": [62.8, None, None]
})
buf = BytesIO()
with pd.ExcelWriter(buf, engine="openpyxl") as writer:
    example_cvas.to_excel(writer, sheet_name="CVAS", index=False)
buf.seek(0)
st.download_button("📥 Download CVAS Excel Template", buf,
                   file_name="CVAS_template.xlsx",
                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ----------------------------
# Step 1: Upload CVAS
# ----------------------------
uploaded_file = st.file_uploader("Upload your CVAS Excel file", type=["xlsx"])

if uploaded_file is not None:
    cvas_df = pd.read_excel(uploaded_file)
    st.subheader("Preview of Uploaded File")
    st.dataframe(cvas_df.head())

    # Treatments
    trial_id = st.text_input("Trial ID", value="MSU41_24ES3")
    num_treatments = st.number_input("Number of Experimental Treatments", min_value=1, step=1, value=2)
    include_pre = st.checkbox("Include PRE-experimental treatment?")
    include_post = st.checkbox("Include POST-experimental treatment?")
    treatment_names = {}
    if include_pre:
        treatment_names["PRE"] = "PRE"
    for t in range(1, num_treatments + 1):
        treatment_names[f"T{t}"] = st.text_input(f"Name for Treatment {t}", value=f"T{t}", key=f"treat_{t}")
    if include_post:
        treatment_names["POST"] = "POST"

    # Feeds + inclusions + NASEM match
    st.subheader("Define feeds, inclusions, and NASEM matches")
    feed_mapping = {}
    use_nasem = st.checkbox("Use NASEM 2021 feed library values?")
    nasem_df = nasem_df_builtin.copy()
    if use_nasem:
        override = st.file_uploader("Upload custom NASEM library", type=["csv"])
        if override:
            nasem_df = pd.read_csv(override)
            nasem_df["Feed Name"] = nasem_df["Feed Name"].astype(str)
            nasem_df["Feed Name (Original)"] = nasem_df["Feed Name"]
            nasem_df["Feed Name (Clean)"] = nasem_df["Feed Name"].apply(norm_text)

    for _, row in cvas_df[["desc_1", "feedtype"]].drop_duplicates().iterrows():
        feed = row["desc_1"]
        feedtype = str(row["feedtype"]).upper()
        fname = str(feed).upper()
        if "FORAGE" in feedtype or "SILAGE" in fname or "HAYLAGE" in fname:
            default_type, default_method = "FORAGE", "NIR, starch, NDFD48"
        elif "MIX" in fname or "BASE" in fname:
            default_type, default_method = "PREMIX", "WC"
        else:
            default_type, default_method = "GRAIN", "NIR, starch"

        with st.expander(f"Feed: {feed}", expanded=False):
            lab = st.text_input(f"Lab for {feed}", "Cumberland Valley Analytical Services", key=f"lab_{feed}")
            type_val = st.text_input(f"Type for {feed}", default_type, key=f"type_{feed}")
            method = st.text_input(f"Method for {feed}", default_method, key=f"method_{feed}")
            inclusions = {t: st.number_input(f"Inclusion % {feed} ({tname})", 0.0, 100.0, 0.0, 0.1,
                                             key=f"inc_{feed}_{t}") for t, tname in treatment_names.items()}

            chosen_feed = "(None)"
            if use_nasem and "Feed Name (Original)" in nasem_df.columns:
                options = sorted(nasem_df["Feed Name (Original)"].dropna().unique().tolist()) + ["(None)"]
                chosen_feed = st.selectbox(f"Select NASEM match for {feed}", options,
                                           index=len(options)-1, key=f"nasem_{feed}")
            feed_mapping[feed] = {"LAB": lab, "TYPE": type_val, "METHOD": method,
                                  "Inclusions": inclusions, "NASEM_match": chosen_feed}

    # Date ranges
    st.subheader("Define date ranges per treatment")
    if "date_ranges" not in st.session_state:
        st.session_state.date_ranges = {t: [] for t in treatment_names}

    for t, tname in treatment_names.items():
        with st.expander(f"Diet: {tname}", expanded=False):
            ranges = st.session_state.date_ranges[t]
            for i, r in enumerate(ranges):
                c1, c2, c3 = st.columns([1, 1, 1])
                with c1:
                    start = st.date_input(f"Start {tname} {i+1}", r.get("start", datetime.today().date()), key=f"{t}_start_{i}")
                with c2:
                    end = st.date_input(f"End {tname} {i+1}", r.get("end", datetime.today().date()), key=f"{t}_end_{i}")
                with c3:
                    dm = st.number_input(f"TMR DM% {tname} {i+1}", 0.0, 100.0, r.get("dm", 50.0), 0.1, key=f"{t}_dm_{i}")
                r.update({"start": start, "end": end, "dm": dm})
                if st.button(f"Delete range {i+1} ({tname})", key=f"del_{t}_{i}"):
                    ranges.pop(i)
                    st.rerun()
            if st.button(f"Add range for {tname}", key=f"add_{t}"):
                last_end = ranges[-1]["end"] if ranges else datetime.today().date()
                new_start = last_end + timedelta(days=1)
                ranges.append({"start": new_start, "end": new_start + timedelta(days=7), "dm": 50.0})
                st.rerun()

    # Generate
    if st.button("Create FICOMP & DIETCOMP Sheets"):
        # Calc derived values
        cvas_df["Trial_ID"] = trial_id
        cvas_df["FI"] = cvas_df["desc_1"]
        cvas_df["LAB"] = cvas_df["desc_1"].map(lambda x: feed_mapping[x]["LAB"])
        cvas_df["TYPE"] = cvas_df["desc_1"].map(lambda x: feed_mapping[x]["TYPE"])
        cvas_df["METHOD"] = cvas_df["desc_1"].map(lambda x: feed_mapping[x]["METHOD"])
        cvas_df["aNDF"] = cvas_df["NDF"]
        cvas_df["FA"] = cvas_df["TFA"]
        cvas_df["RUP"] = cvas_df["CP"] - cvas_df["RDP"]
        cvas_df["OM"] = 100 - cvas_df["Ash"]
        cvas_df["ForNDF"] = cvas_df.apply(lambda r: r["aNDFom"] if str(r["TYPE"]).upper()=="FORAGE" else 0, axis=1)
        for col in ["NDFD30","NDFD48","ADF","Lignin","Starch"]:
            if col not in cvas_df.columns: cvas_df[col]=None

        # NASEM fallback
        nasem_map={"aNDF":"Feed NDF","NDFD48":"Feed DNDF48_NDF","RUP":"Feed RUP_base","FA":"Feed FA"}
        for col,nasem_col in nasem_map.items():
            flag=f"{col}_Flag"; cvas_df[flag]="ACTUAL"
            for feed in cvas_df["desc_1"]:
                match=feed_mapping[feed]["NASEM_match"]
                if match and match!="(None)" and nasem_col in nasem_df.columns:
                    if pd.isna(cvas_df.loc[cvas_df["desc_1"]==feed,col]).all():
                        row=nasem_df[nasem_df["Feed Name (Original)"]==match]
                        if not row.empty:
                            cvas_df.loc[cvas_df["desc_1"]==feed,col]=row[nasem_col].values[0]
                            cvas_df.loc[cvas_df["desc_1"]==feed,flag]="NASEM"

        # FICOMP
        ficomp_cols=["Trial_ID","FI","LAB","TYPE","METHOD","OM",
                     "aNDF","aNDF_Flag","aNDFom","ForNDF","NDFD30",
                     "NDFD48","NDFD48_Flag","ADF","Lignin","Starch",
                     "CP","RUP","RUP_Flag","FA","FA_Flag","Ash"]
        ficomp_df=cvas_df[ficomp_cols]

        # DIETCOMP
        diet_rows=[]
        for t,tname in treatment_names.items():
            for rng in st.session_state.date_ranges[t]:
                for d in pd.date_range(rng["start"],rng["end"]):
                    row={"Trial_ID":trial_id,"Diet":tname,"Date":d,"TMR_DM":rng["dm"]}
                    for col in ["OM","aNDF","aNDFom","ForNDF","NDFD30","NDFD48","ADF","Lignin","Starch","CP","RUP","FA","Ash"]:
                        val=0
                        for feed in feed_mapping:
                            incl=feed_mapping[feed]["Inclusions"].get(t,0)/100.0
                            feed_val=cvas_df.loc[cvas_df["desc_1"]==feed,col].values
                            if len(feed_val)>0 and not pd.isna(feed_val[0]): val+=incl*feed_val[0]
                        row[col]=val
                    diet_rows.append(row)
        dietcomp_df=pd.DataFrame(diet_rows)

        # Key
        key_df=pd.DataFrame({
            "Variable":ficomp_df.columns,
            "Definition":["See documentation"]*len(ficomp_df.columns)
        })

        # Save
        fname=f"{trial_id}_DietComp.xlsx"
        with pd.ExcelWriter(fname,engine="openpyxl") as w:
            ficomp_df.to_excel(w,"FICOMP",index=False)
            dietcomp_df.to_excel(w,"DIETCOMP",index=False)
            key_df.to_excel(w,"KEY",index=False)
        auto_fit_and_format(fname)

        st.success("✅ FICOMP, DIETCOMP & Key generated!")
        st.dataframe(ficomp_df.head())
        st.dataframe(dietcomp_df.head())
        with open(fname,"rb") as f:
            st.download_button("📥 Download Excel",f,file_name=fname)
